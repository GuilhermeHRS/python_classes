import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import os


def celsiusToFahrenheit():
    input_text = nome_entry.get()

    if not input_text:
        messagebox.showerror("Erro", "Digite um valor válido em Celsius")
        return

    try:
        celsius = float(input_text)
    except ValueError:
        messagebox.showerror("Erro", "Digite um valor numérico válido em Celsius")
        return

    fahrenheit = (celsius * 1.8) + 32
    msg_label.config(text=f"{celsius} Celsius são {fahrenheit:.2f} Fahrenheit")
    nome_entry.delete(0, tk.END)

    # Armazenar os dados em uma planilha -- insira o caminho da pasta  
    filename = "projects/conversoes.xlsx"

    # Verificar se o arquivo já existe
    if not os.path.isfile(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Data e Hora", "Celsius", "Fahrenheit"])
    else:
        wb = load_workbook(filename)
        sheet = wb.active

    sheet.append([datetime.now(), celsius, fahrenheit])
    wb.save(filename)

# Criar a window principal
window = tk.Tk()
window.title("Conversor Celsius para Fahrenheit")

# Aumentar o tamanho da interface
width = 400
height = 200
window.geometry(f"{width}x{height}")

# Criar um rótulo
msg_label = tk.Label(window, text="Digite os graus em Celsius")
msg_label.pack(pady=20)

# Criar um campo de entrada
nome_entry = tk.Entry(window)
nome_entry.pack()

# Criar um botão
botao = tk.Button(window, text="Converter", command=celsiusToFahrenheit)
botao.pack()

window.mainloop()

